import openpyxl

import pandas

from openpyxl.styles import PatternFill, Font
from dateutil.relativedelta import relativedelta
from datetime import datetime
from pandas.io.excel import ExcelWriter

csv_files=['output.csv'] #Needed for below function
with ExcelWriter('output.xlsx') as ew: #TAKES IN THE OUTPUT CSV TO TURN IT INTO AN EXCEL FILE!
    for csv_file in csv_files:
        pandas.read_csv(csv_file).to_excel(ew, sheet_name=csv_file)


#THIS IS THE SECOND SCRIPT, ENTER THE NAME OF YOUR EXCEL SHEET AND IT'S SHEET NAME BELOW
#DO NOT FORGET TO READ THE INSTRUCTIONS! IF SCRIPT DOESN'T WORK, IT'S YOUR FAULT!
#IF A FAULTY EXCEL FILE IS CREATED, DELETE IT AND RUN THE SCRIPT AFTER CORRECTING THE CODE
dt = pandas.read_excel("output.xlsx", sheet_name="output.csv") #Make sure you input the correct file and sheet name!
wb = openpyxl.load_workbook("output.xlsx")
ws = wb['output.csv'] #Name of the working sheet

try: #This changes the text color of colorCode columns based on it's value, paints it white if value is None.
    for cell in ws['D']:#MAKE SURE YOU USE THE COLUMN ID OF COLOR CODES FROM THE output.csv!!!
                        # REMEMBER THAT RNR (always the first) HAS THE COLUMN B!!!
        try:
            cell.font = Font(color=f'{cell.value}')
        except:
            cell.font = Font(color='ffffff')
except:
    pass

dates=[]#This is needed for the below function to work.
for cell in ws['E']:  # MAKE SURE YOU USE THE COLUMN ID OF HU FROM THE output.csv!!!
                      # REMEMBER THAT RNR (always the first) HAS THE COLUMN B!!!
        dates.append(cell.value)

today = pandas.to_datetime('today').normalize()

for date in dates:  #THIS COLORS THE DATES ON HOW OLD THEY ARE, DUPLICATE DATES AREN'T COLORED.
    current_index =dates.index(date)+1

    try:
        date = datetime.strptime(date, '%Y-%m-%d')
        date= date + relativedelta(days=+14)
        str_d1 = '2021/10/20'
        str_d2 = '2022/2/20'

        d2 = datetime.strptime(str_d2, "%Y/%m/%d")
        delta = today - date
        if delta.days>=360: #If older than 12 monts
            print(f'Difference is {delta.days},more than a year past!')
            for col, val in enumerate(dates, start=1):
                ws.cell(row=current_index, column=col).fill = PatternFill(start_color='b30000', end_color='b30000',fill_type="solid")

        elif delta.days <= 90: #If not older than 3 months
            print(f'Difference is {delta.days},not older than 3 months')
            for col, val in enumerate(dates, start=1):
                ws.cell(row=current_index, column=col).fill = PatternFill(start_color='007500', end_color='007500',fill_type="solid")

        elif 90< delta.days < 365:#If not oldar than a year, but older than 3 months
            print(f'Difference is {delta.days},not older than a year')
            for col, val in enumerate(dates, start=1):
                ws.cell(row=current_index, column=col).fill = PatternFill(start_color='ffa500', end_color='ffa500',fill_type="solid")
        else:
            print(f'Difference is {delta.days}')
    except:
        print("Date not given")

count=0
for cell in ws['F']: #THIS UPDATES THE DATES
    cell.value=dates[count]
    count+=1

wb.save("output.xlsx")

